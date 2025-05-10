# ✅ 1. coupang_crawler.py (터미널 실행용 - 시트 분할, 수식 적용, 이미지 포함 최종 버전)
import pandas as pd
import time
from urllib.parse import quote
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook


def crawl(keyword):
    encoded_keyword = quote(keyword)
    data = []

    options = uc.ChromeOptions()
    options.add_argument("--disable-gpu")
    options.add_argument("user-agent=Mozilla/5.0")

    service = Service("chromedriver.exe")
    driver = uc.Chrome(service=service, options=options)

    for page in range(1, 2):
        url = f"https://www.coupang.com/np/search?q={encoded_keyword}&channel=user&listSize=72&page={page}"
        print(f"페이지 접속 중: {url}")
        driver.get(url)
        time.sleep(2)

        items = driver.find_elements(By.CLASS_NAME, "search-product")

        for item in items:
            if "ad-badge" in item.get_attribute("innerHTML"):
                continue

            try:
                name = item.find_element(By.CLASS_NAME, "name").text.strip()
            except:
                name = "상품명 없음"

            try:
                link = item.find_element(By.TAG_NAME, "a").get_attribute("href")
            except:
                link = "링크 없음"

            try:
                origin_price = item.find_element(By.CLASS_NAME, "base-price").text
            except:
                origin_price = "없음"

            try:
                price = item.find_element(By.CLASS_NAME, "price-value").text
            except:
                price = "없음"

            try:
                card_discount = item.find_element(By.CLASS_NAME, "ccid-txt").text
            except:
                card_discount = "없음"

            try:
                point = item.find_element(By.CLASS_NAME, "reward-cash-txt").text
            except:
                point = "없음"

            try:
                img = item.find_element(By.CLASS_NAME, "search-product-wrap-img").get_attribute("src")
                img_url = f"https:{img}" if img.startswith("//") else img
                image_formula = f'=IMAGE("{img_url}")'
            except:
                image_formula = "이미지 없음"

            rocket = "로켓배송" if item.find_elements(By.XPATH, './/img[@alt="로켓배송"]') else "일반배송"

            data.append({
                "이미지": image_formula,
                "상품명": name,
                "정가": origin_price,
                "판매가": price,
                "카드할인정보": card_discount,
                "적립금": point,
                "배송유형": rocket,
                "URL": link
            })

        time.sleep(1)

    driver.quit()

    filename = f"{keyword}_크롤링결과.xlsx"
    df = pd.DataFrame(data)
    df['이미지'] = df['이미지'].str.replace('^@', '', regex=True)

    rocket_df = df[df['배송유형'] == '로켓배송']
    normal_df = df[df['배송유형'] != '로켓배송']

    with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='전체상품', index=False)
        rocket_df.to_excel(writer, sheet_name='로켓배송', index=False)
        normal_df.to_excel(writer, sheet_name='일반배송', index=False)

    wb = load_workbook(filename)
    for sheet_name in ['전체상품', '로켓배송', '일반배송']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            image_col_idx = list(df.columns).index("이미지") + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=image_col_idx)
                if isinstance(cell.value, str) and cell.value.startswith("=IMAGE"):
                    cell._value = cell.value
                    cell.data_type = 'f'
    wb.save(filename)
    print(f"✅ 크롤링 완료: {filename}")


if __name__ == "__main__":
    keyword = input("검색 키워드를 입력하세요: ")
    if keyword:
        crawl(keyword)
    else:
        print("⛔ 키워드를 입력해주세요.")
