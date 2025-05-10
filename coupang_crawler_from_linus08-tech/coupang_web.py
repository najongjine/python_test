# ✅ Streamlit 웹 앱용 쿠팡 크롤러 (페이지 선택 + 로켓배송 필터 + 시트 분리 + 썸네일 수식 포함)
import streamlit as st
import pandas as pd
import time
from urllib.parse import quote
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook
import os

# 💻 크롤링 함수
def crawl_coupang(keyword, max_pages=3, rocket_only=False):
    encoded_keyword = quote(keyword)
    chrome_driver_path = "chromedriver.exe"
    service = Service(executable_path=chrome_driver_path)

    options = uc.ChromeOptions()
    options.add_argument("--disable-gpu")
    options.add_argument("user-agent=Mozilla/5.0")

    driver = uc.Chrome(service=service, options=options)
    data = []

    for page in range(1, max_pages + 1):
        url = f"https://www.coupang.com/np/search?q={encoded_keyword}&channel=user&listSize=72&page={page}"
        driver.get(url)
        time.sleep(2)

        items = driver.find_elements(By.CLASS_NAME, "search-product")

        for item in items:
            if "ad-badge" in item.get_attribute("innerHTML"):
                continue

            try:
                name = item.find_element(By.CLASS_NAME, "name").text.strip()
            except:
                name = "상품명 없음"

            try:
                link = item.find_element(By.TAG_NAME, "a").get_attribute("href")
            except:
                link = "링크 없음"

            try:
                origin_price = item.find_element(By.CLASS_NAME, "base-price").text
            except:
                origin_price = "없음"

            try:
                price = item.find_element(By.CLASS_NAME, "price-value").text
            except:
                price = "없음"

            try:
                card_discount = item.find_element(By.CLASS_NAME, "ccid-txt").text
            except:
                card_discount = "없음"

            try:
                point = item.find_element(By.CLASS_NAME, "reward-cash-txt").text
            except:
                point = "없음"

            try:
                img = item.find_element(By.CLASS_NAME, "search-product-wrap-img").get_attribute("src")
                img_url = f"https:{img}" if img.startswith("//") else img
                image_formula = f'=IMAGE("{img_url}")'
            except:
                image_formula = "이미지 없음"

            rocket = "로켓배송" if item.find_elements(By.XPATH, './/img[@alt="로켓배송"]') else "일반배송"

            # ✅ 로켓배송 필터 적용
            if rocket_only and rocket != "로켓배송":
                continue

            data.append({
                "이미지": image_formula,
                "상품명": name,
                "정가": origin_price,
                "판매가": price,
                "카드할인정보": card_discount,
                "적립금": point,
                "배송유형": rocket,
                "URL": link
            })

        time.sleep(1)

    driver.quit()

    df = pd.DataFrame(data)
    df['이미지'] = df['이미지'].str.replace('^@', '', regex=True)
    return df

# 🌐 Streamlit 웹 UI 구성
st.set_page_config(page_title="쿠팡 크롤러", layout="centered")
st.title("🛒 쿠팡 키워드 크롤러")

keyword = st.text_input("🔍 키워드를 입력하세요")
page_count = st.slider("🔢 몇 페이지까지 크롤링할까요?", min_value=1, max_value=10, value=3)
rocket_only = st.checkbox("🚀 로켓배송만 보기")

if st.button("크롤링 시작"):
    if not keyword:
        st.warning("⛔ 키워드를 입력해 주세요.")
    else:
        with st.spinner("쿠팡에서 크롤링 중입니다..."):
            df = crawl_coupang(keyword, max_pages=page_count, rocket_only=rocket_only)
            rocket_df = df[df['배송유형'] == '로켓배송']
            normal_df = df[df['배송유형'] != '로켓배송']
            filename = f"{keyword}_크롤링결과.xlsx"

            with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, sheet_name="전체상품", index=False)
                rocket_df.to_excel(writer, sheet_name="로켓배송", index=False)
                normal_df.to_excel(writer, sheet_name="일반배송", index=False)

            wb = load_workbook(filename)
            for sheet in ['전체상품', '로켓배송', '일반배송']:
                ws = wb[sheet]
                col_idx = list(df.columns).index("이미지") + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if isinstance(cell.value, str) and cell.value.startswith("=IMAGE"):
                        cell._value = cell.value
                        cell.data_type = 'f'
            wb.save(filename)

        st.success("✅ 크롤링 완료! 아래에서 파일을 다운로드하세요.")
        with open(filename, "rb") as f:
            st.download_button("📥 엑셀 다운로드", f, file_name=filename)
        os.remove(filename)
